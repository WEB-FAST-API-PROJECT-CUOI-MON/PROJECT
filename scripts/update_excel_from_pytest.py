"""Chạy pytest rồi tự động ghi kết quả vào cột "Kết quả thực tế" / "Pass/Fail" trong
docs/Checklist_Test_API.xlsx.

Cách dùng:
    venv\\Scripts\\python scripts\\update_excel_from_pytest.py

Tuỳ chọn:
    --skip-run          Không chạy lại pytest, dùng report JSON đã có sẵn (--report).
    --report PATH        Đường dẫn file JSON report (mặc định: .pytest_report.json ở gốc repo).
    --workbook PATH       Đường dẫn file Excel cần cập nhật (mặc định: docs/Checklist_Test_API.xlsx).
    --tester NAME         Tên ghi vào cột "Người test" (mặc định: "pytest (tự động)").

LƯU Ý QUAN TRỌNG — đọc trước khi dùng:
1. File checklist Excel được thiết kế để test THỦ CÔNG qua Swagger/Postman (xem sheet
   "Tổng quan"), nên nó có nhiều dòng ("Mã TC") chi tiết hơn số lượng test pytest hiện có
   (~126 dòng Mã TC so với 66 test tự động). Những dòng KHÔNG có test pytest tương ứng
   (bảng MAPPING bên dưới không liệt kê) sẽ bị BỎ QUA — vẫn cần test tay bình thường.
2. Một số test pytest xác minh nhiều tình huống trong cùng 1 hàm (vd. vừa test case đúng
   vừa test case lỗi liền nhau). Khi đó nhiều Mã TC cùng trỏ về 1 test — nếu test đó FAIL,
   TẤT CẢ Mã TC liên quan sẽ bị đánh dấu Fail (script không biết chính xác assertion nào
   trong hàm bị lỗi).
3. Đóng file Excel trước khi chạy script — Excel khoá file khi đang mở nên ghi đè sẽ lỗi.
4. Khi thêm test pytest mới, nhớ bổ sung vào MAPPING bên dưới nếu muốn nó tự động điền vào
   đúng dòng Mã TC tương ứng.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from datetime import date
from pathlib import Path

import openpyxl

# Console Windows mặc định dùng codepage (vd. cp1252), không encode được tiếng Việt có dấu
# hay các ký tự như →/✔/⚠ — ép stdout/stderr sang UTF-8 để print() không bị crash.
for _stream in (sys.stdout, sys.stderr):
    if _stream and _stream.encoding and _stream.encoding.lower() != "utf-8":
        _stream.reconfigure(encoding="utf-8", errors="replace")

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = PROJECT_ROOT / "docs" / "Checklist_Test_API.xlsx"
DEFAULT_REPORT = PROJECT_ROOT / ".pytest_report.json"

# Các sheet chỉ chứa hướng dẫn/dữ liệu mẫu, không có checklist test case.
SKIP_SHEETS = {"Tổng quan", "Dữ liệu mẫu"}

# Cột trong mỗi sheet checklist (1-indexed, theo openpyxl).
COL_MA_TC = 7
COL_EXPECTED_STATUS = 13
COL_ACTUAL_RESULT = 15
COL_PASS_FAIL = 16
COL_TESTER = 17
COL_TEST_DATE = 18

# Mã TC -> tên hàm test trong tests/ (tên hàm là duy nhất trong toàn bộ test suite).
# Xem docstring ở đầu file: chỉ những Mã TC liệt kê ở đây mới được tự động điền.
MAPPING: dict[str, str] = {
    # --- Health ---
    "HC-01": "test_health",
    # --- Auth: register/login/refresh ---
    "REG-01": "test_register_success",
    "REG-02": "test_register_duplicate_email",
    "REG-04": "test_register_password_too_short",
    "LOG-01": "test_login_success_returns_access_and_refresh_token",
    "LOG-02": "test_login_wrong_password",
    "LOG-04": "test_login_inactive_account",
    "LOG-06": "test_login_rate_limit_blocks_after_max_attempts",
    "REF-01": "test_refresh_token_issues_new_access_token",
    "REF-04": "test_refresh_rejects_access_token",
    # --- Users ---
    "USR-01": "test_get_current_user",
    "USR-02": "test_get_current_user_without_token",
    "USR-03": "test_get_current_user_expired_token",
    "USR-04": "test_get_current_user_invalid_token",
    "USR-05": "test_list_users_as_admin_with_search_and_status_filter",
    "USR-06": "test_list_users_as_admin_with_search_and_status_filter",
    "USR-07": "test_list_users_as_admin_with_search_and_status_filter",
    "USR-08": "test_list_users_requires_admin",
    # --- Construction Sites ---
    "SITE-01": "test_create_site_success_and_becomes_owner",
    "SITE-03": "test_create_site_name_blank_rejected",
    "SITE-04": "test_create_site_requires_auth",
    "SITE-05": "test_list_sites_only_returns_owner_or_member",
    "SITE-06": "test_list_sites_search_by_name",
    "SITE-09": "test_get_site_not_found",
    "SITE-10": "test_get_site_requires_membership",
    "SITE-11": "test_delete_site_soft_delete_hides_it",
    "SITE-12": "test_update_site_owner_only",
    "SITE-13": "test_update_site_via_put_same_as_patch",
    "SITE-14": "test_update_site_owner_only",
    "SITE-17": "test_delete_site_soft_delete_hides_it",
    "SITE-18": "test_delete_site_non_owner_forbidden",
    "SM-03": "test_add_member_success_and_duplicate_rejected",
    "SM-05": "test_add_member_success_and_duplicate_rejected",
    "SM-06": "test_add_member_user_not_found",
    "SM-07": "test_add_member_non_owner_forbidden",
    "SM-09": "test_remove_member_success",
    "SM-10": "test_cannot_remove_last_owner",
    "SM-11": "test_remove_member_not_found",
    # --- Teams ---
    "TEAM-01": "test_create_team_owner_only",
    "TEAM-02": "test_create_team_owner_only",
    "TEAM-06": "test_team_list_requires_site_membership",
    "TEAM-07": "test_get_team_detail_requires_membership",
    "TEAM-08": "test_get_team_not_found",
    "TEAM-09": "test_update_team_owner_only",
    "TEAM-10": "test_update_team_owner_only",
    "TEAM-11": "test_delete_team_owner_only_and_unassigns_work_items",
    "TEAM-12": "test_delete_team_owner_only_and_unassigns_work_items",
    "TM-01": "test_list_team_members",
    "TM-03": "test_add_team_member_requires_site_membership",
    "TM-04": "test_add_team_member_duplicate_rejected",
    "TM-06": "test_remove_team_member_owner_only_and_not_found",
    "TM-07": "test_remove_team_member_owner_only_and_not_found",
    # --- Work Items ---
    "WI-01": "test_create_work_item_with_valid_assignee_team",
    "WI-03": "test_create_work_item_assignee_team_must_belong_to_site",
    "WI-04": "test_create_work_item_requires_membership",
    "WI-07": "test_list_work_items_scoped_to_site",
    "WI-08": "test_list_work_items_filter_and_search",
    "WI-09": "test_list_work_items_filter_and_search",
    "WI-10": "test_list_work_items_pagination",
    "WI-13": "test_list_work_items_scoped_to_site",
    "WI-14": "test_get_work_item_detail_requires_membership",
    "WI-15": "test_get_work_item_not_found",
    "WI-16": "test_get_work_item_detail_requires_membership",
    "WI-17": "test_update_work_item_owner_can_edit_any_field",
    "WI-18": "test_update_work_item_assignee_team_member_can_only_update_status_description",
    "WI-19": "test_update_work_item_assignee_team_member_can_only_update_status_description",
    "WI-20": "test_update_work_item_unrelated_member_forbidden",
    "WI-23": "test_delete_work_item_owner_only",
    "WI-24": "test_delete_work_item_owner_only",
    # --- Comments ---
    "CMT-01": "test_comment_create_and_list_requires_membership",
    "CMT-02": "test_comment_create_and_list_requires_membership",
    "CMT-03": "test_comment_create_and_list_requires_membership",
    "CMT-07": "test_comment_create_and_list_requires_membership",
    # --- Attachments ---
    "ATT-02": "test_upload_attachment_success_and_list",
    "ATT-03": "test_upload_attachment_rejects_invalid_type",
    "ATT-07": "test_upload_attachment_requires_membership",
    "ATT-08": "test_upload_attachment_success_and_list",
}

_ASSERT_STATUS_RE = re.compile(r"assert\s+(\d{3})\s*==\s*(\d{3})")


def run_pytest(report_path: Path, extra_args: list[str]) -> int:
    cmd = [
        sys.executable,
        "-m",
        "pytest",
        "-q",
        f"--json-report-file={report_path}",
        "--json-report",
        *extra_args,
    ]
    print("→ Chạy:", " ".join(cmd))
    result = subprocess.run(cmd, cwd=PROJECT_ROOT)
    return result.returncode


def load_report(report_path: Path) -> dict[str, dict]:
    """Trả về {tên_hàm_test: {"outcome": ..., "actual_status": int|None}}."""
    data = json.loads(report_path.read_text(encoding="utf-8"))
    by_name: dict[str, dict] = {}
    for test in data.get("tests", []):
        node_id = test["nodeid"]
        name = node_id.split("::")[-1]
        outcome = test["outcome"]  # "passed" / "failed" / "skipped" / ...
        actual_status = None
        if outcome == "failed":
            call = test.get("call") or {}
            crash_msg = (call.get("crash") or {}).get("message", "")
            m = _ASSERT_STATUS_RE.search(crash_msg)
            if m:
                actual_status = int(m.group(1))
        by_name[name] = {"outcome": outcome, "actual_status": actual_status}
    return by_name


def find_header_row(ws) -> int | None:
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, COL_MA_TC).value == "Mã TC":
            return r
    return None


def update_workbook(workbook_path: Path, results_by_name: dict[str, dict], tester: str) -> tuple[int, list[str]]:
    wb = openpyxl.load_workbook(workbook_path)
    updated = 0
    unmapped_with_no_result: list[str] = []
    today = date.today()

    for ws in wb.worksheets:
        if ws.title in SKIP_SHEETS:
            continue
        header_row = find_header_row(ws)
        if header_row is None:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            ma_tc = ws.cell(r, COL_MA_TC).value
            if not ma_tc:
                continue
            test_name = MAPPING.get(ma_tc)
            if test_name is None:
                continue  # không có test tự động cho case này -> để test thủ công
            result = results_by_name.get(test_name)
            if result is None:
                unmapped_with_no_result.append(f"{ma_tc} -> {test_name} (không có trong report, có thể đã bị lọc/deselect)")
                continue

            expected_status = ws.cell(r, COL_EXPECTED_STATUS).value
            if result["outcome"] == "passed":
                actual_text = f"{expected_status} — khớp kỳ vọng (pytest: {test_name}, PASS)"
                pass_fail = "Pass"
            elif result["outcome"] == "failed":
                if result["actual_status"] is not None:
                    actual_text = (
                        f"{result['actual_status']} — KHÁC kỳ vọng {expected_status} "
                        f"(pytest: {test_name}, FAILED)"
                    )
                else:
                    actual_text = f"FAILED (pytest: {test_name}) — xem report.json để biết chi tiết"
                pass_fail = "Fail"
            else:  # skipped / error / ...
                actual_text = f"{result['outcome'].upper()} (pytest: {test_name})"
                pass_fail = "Blocked"

            ws.cell(r, COL_ACTUAL_RESULT).value = actual_text
            ws.cell(r, COL_PASS_FAIL).value = pass_fail
            ws.cell(r, COL_TESTER).value = tester
            ws.cell(r, COL_TEST_DATE).value = today
            updated += 1

    try:
        wb.save(workbook_path)
    except PermissionError as exc:
        raise SystemExit(
            f"Không ghi được vào '{workbook_path}' — file đang mở trong Excel. "
            "Hãy đóng file rồi chạy lại script."
        ) from exc

    return updated, unmapped_with_no_result


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--skip-run", action="store_true", help="Dùng report JSON có sẵn, không chạy lại pytest")
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--tester", default="pytest (tự động)")
    parser.add_argument("pytest_args", nargs="*", help="Tham số truyền thêm cho pytest (vd. tests/test_auth.py)")
    args = parser.parse_args()

    if not args.skip_run:
        run_pytest(args.report, args.pytest_args)
    elif not args.report.exists():
        raise SystemExit(f"Không tìm thấy report '{args.report}'. Bỏ --skip-run để chạy pytest trước.")

    results_by_name = load_report(args.report)
    updated, missing = update_workbook(args.workbook, results_by_name, args.tester)

    print(f"\n✔ Đã cập nhật {updated} dòng trong '{args.workbook.name}'.")
    if missing:
        print(f"⚠ {len(missing)} Mã TC có mapping nhưng không thấy trong report pytest lần này:")
        for line in missing:
            print("   -", line)
    print(
        f"ℹ Có {len(MAPPING)} Mã TC được tự động hoá qua pytest. Các Mã TC còn lại trong "
        "checklist vẫn cần test thủ công qua Swagger/Postman và tự điền kết quả."
    )


if __name__ == "__main__":
    main()
